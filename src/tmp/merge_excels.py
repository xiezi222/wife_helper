#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

out_file_path = os.path.join(os.getcwd(), "合并.xlsx")
cell_colors = ["00FF7F", "FFFF00", "00BFFF", "FF00FF", "FFA500", "008B8B"]  # 绿色，黄色。蓝色。粉色.橙色,深青色
current_color = ["000000"]


def remove_old_out_file(path):
    if os.path.exists(path):
        print("删除之前的输出文件")
        os.remove(path)


def get_all_excel_file_paths():
    print("开始查找excel文件(.xlsx)……")
    path = os.getcwd()
    file_paths = []
    for file_name in os.listdir(path):
        if file_name.endswith("xlsx") and not (file_name.startswith("~") or file_name.startswith(".")):
            print("找到excel文件：", file_name)
            file_paths.append(os.path.join(path, file_name))
    if len(file_paths) == 0:
        print("没有找到excel文件")
    return file_paths


def get_custom_titles_map_from_sheet(sheet):
    custom_title_map = {}
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in row:
            custom_title = cell.value
            if custom_title is not None:
                custom_title_map[custom_title] = cell.column
    return custom_title_map


def merge_sheet(source_sheet, target_sheet):
    custom_titles_map = get_custom_titles_map_from_sheet(target_sheet)
    target_sheet_row_count = target_sheet.max_row

    for columns in source_sheet.columns:
        top_cell = columns[0]
        custom_title = top_cell.value

        if custom_title not in custom_titles_map.keys():
            column = len(custom_titles_map.keys()) + 1
            custom_titles_map[custom_title] = column
            target_sheet.cell(row=1, column=column, value=custom_title)
        target_column = custom_titles_map[custom_title]

        for cell in columns:
            if cell.row == 1:
                continue
            new_cell = target_sheet.cell(row=target_sheet_row_count + cell.row - 1,
                                         column=target_column,
                                         value=cell.value)
            # 设置单元格填充色
            new_cell.fill = PatternFill("solid", fgColor=current_color)


def start_merge():
    remove_old_out_file(out_file_path)
    new_workbook = Workbook()

    in_file_paths = get_all_excel_file_paths()
    for file_path in in_file_paths:
        print("开始合并文件:", os.path.basename(file_path))
        global current_color
        current_color = cell_colors[in_file_paths.index(file_path) % len(in_file_paths)]
        workbook = load_workbook(file_path)
        for sheet in workbook:
            sheet_name = sheet.title
            if sheet_name not in new_workbook.sheetnames:
                new_workbook.create_sheet(title=sheet_name)
            merge_sheet(workbook[sheet_name], new_workbook[sheet_name])
            print("sheet 合并完成:", sheet_name)
        workbook.close()
        print("文件合并完成:", os.path.basename(file_path))

    new_workbook.save(out_file_path)
    new_workbook.close()


if __name__ == '__main__':
    print("开始执行任务……")
    start_time = datetime.now()
    start_merge()
    delta = (datetime.now() - start_time).seconds
    print("任务执行完毕, 共耗时:", int(delta / 60), "分", delta % 60, "秒")
