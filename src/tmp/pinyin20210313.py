#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import pypinyin
import os
import time
import re

from pypinyin import Style
from datetime import datetime
from openpyxl import Workbook, load_workbook, cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment


in_path = os.path.join(os.getcwd(),"合并.xlsx")
out_path = os.path.join(os.getcwd(), "../xlsx/转置.xlsx")

title_map = {}

def log_error(str):
    print("\033[0;31;43m警告:"+ str + "\033[0m")


def remove_old_out_file(path):
    if os.path.exists(path):
        print("删除之前的输出文件")
        os.remove(path)

def get_first_letter(string):

    letter = pypinyin.slug(string, style=Style.FIRST_LETTER, separator='')
    letter = letter.strip()#去除首尾空格
    letter = re.sub(r"[^\w+$]", "_", letter)
    letter = letter.strip('_')#去除首尾下划线
    return letter

def transform(source_sheet, target_sheet):
    rows = source_sheet.iter_rows(min_col=1, max_col=source_sheet.max_column, min_row=1,max_row=1)
    for row in rows:
        for cell in row:
            value = cell.value
            if value == None:
                log_error(str.format("没有列名， 位置：{:s}:{:d}".format(cell.column_letter,cell.row)))
                continue

            letter = ""
            if value in title_map.keys():
                letter = title_map[value]
            else:
                letter = get_first_letter(value)
                title_map[value] = letter
            target_sheet.cell(row=cell.column, column=1, value=value)
            target_sheet.cell(row=cell.column, column=2, value=letter)

def start():
    remove_old_out_file(out_path)
    new_workbook = Workbook() 
    if not os.path.exists(in_path):
        print("没有输入文件：",in_path)
        return

    workbook = load_workbook(in_path)
    for sheet in workbook:
        sheet_name = sheet.title
        print(sheet_name)
        target_sheet = new_workbook.create_sheet(title=sheet_name)
        transform(workbook[sheet_name], target_sheet)
        print("sheet 转置完成:",sheet_name)
    workbook.close()
    new_workbook.save(out_path)
    new_workbook.close()

if __name__ == '__main__':
    print ("开始执行任务……")
    start_time = datetime.now()
    start()
    delta = (datetime.now() - start_time).seconds
    print ("任务执行完毕, 共耗时:", delta / 60, "分", delta % 60, "秒")
    

