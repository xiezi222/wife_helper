#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import os
from base_script import BaseScript
from openpyxl import Workbook, load_workbook


class ExcelToSql(BaseScript):

    def __init__(self, excel_path, sql_path):
        self.excel_path = excel_path
        self.sql_path = sql_path
        self.out_path = os.path.join(os.path.dirname(excel_path), "sql语句.xlsx")

    def remove_old_out_file(self):
        if os.path.exists(self.out_path):
            print("删除之前的输出文件")
            os.remove(self.out_path)

    def merge_sheet(self, source_sheet, target_sheet):
        model_file_path = os.path.join(self.sql_path, source_sheet.title+".txt")
        if not os.path.exists(model_file_path):
            return

        model_file = open(model_file_path, 'r')
        sql = model_file.read()
        for index, rows in enumerate(source_sheet.rows):
            if index == 0:
                continue
            new_sql = sql
            for cell in rows:
                if cell.value is not None:
                    key = "cell" + str(cell.column)
                    new_sql = new_sql.replace(key, cell.value, 1)
                else:
                    print("空内容  位置：列：" + cell.column + "-行：" + cell.row)
                    exit()
            target_sheet.cell(row=index, column=1, value=new_sql)
        model_file.close()

    def executive(self):
        self.remove_old_out_file()
        new_workbook = Workbook()

        print("开始处理文件:", self.excel_path)
        workbook = load_workbook(self.excel_path)
        for sheet in workbook:
            sheet_name = sheet.title
            if sheet_name not in new_workbook.sheetnames:
                new_workbook.create_sheet(title=sheet_name)
            self.merge_sheet(workbook[sheet_name], new_workbook[sheet_name])
            print("sheet 处理完成:", sheet_name)
        workbook.close()
        print("文件处理完成:", self.excel_path)

        new_workbook.save(self.out_path)
        new_workbook.close()
